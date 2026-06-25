from __future__ import annotations

import io
import logging
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from telethon.tl import functions

from ..core.config import settings
from ..core.database import db
from ..core.security import current_user_id
from ..helpers.telegram import (
    _account_session_lock,
    _open_account_client_for_user,
    _resolve_entity,
    _safe_disconnect,

)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="")


@router.get("/api/v1/scrape-group/members")
async def scrape_group_members(
    account_id: str,
    group_id: str,
    user_id: str = Depends(current_user_id),
) -> dict[str, Any]:
    """Scrape members from a group and return as JSON."""
    members = []
    lock = _account_session_lock(account_id)
    async with lock:
        _, client = await _open_account_client_for_user(user_id, account_id)
        try:
            try:
                entity = await _resolve_entity(client, group_id)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Could not resolve group: {str(e)}")

            try:
                async for participant in client.iter_participants(entity):
                    if getattr(participant, "bot", False) or getattr(participant, "deleted", False):
                        continue
                    username = getattr(participant, "username", None)
                    first_name = getattr(participant, "first_name", None) or ""
                    last_name = getattr(participant, "last_name", None) or ""
                    full_name = f"{first_name} {last_name}".strip()
                    phone = getattr(participant, "phone", None)

                    access_hash = getattr(participant, "access_hash", None)
                    members.append({
                        "user_id": str(participant.id),
                        "username": username or "",
                        "full_name": full_name or str(participant.id),
                        "phone": phone or "",
                        "access_hash": str(access_hash) if access_hash is not None else "",
                    })
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Failed to scrape members: {str(e)}")
        finally:
            await _safe_disconnect(client)

    return {"members": members, "count": len(members)}


@router.get("/api/v1/scrape/contacts")
async def scrape_contacts(
    account_id: str,
    user_id: str = Depends(current_user_id),
) -> dict[str, Any]:
    """Scrape all contacts of a Telegram account."""
    members = []
    lock = _account_session_lock(account_id)
    async with lock:
        _, client = await _open_account_client_for_user(user_id, account_id)
        try:
            result = await client(functions.contacts.GetContactsRequest(0))
            contacts = result.users if result else []
            for c in contacts:
                if getattr(c, "bot", False) or getattr(c, "deleted", False):
                    continue
                first = getattr(c, "first_name", None) or ""
                last = getattr(c, "last_name", None) or ""
                full_name = f"{first} {last}".strip()
                username = getattr(c, "username", None)
                phone = getattr(c, "phone", None)
                access_hash = getattr(c, "access_hash", None)
                members.append({
                    "user_id": str(c.id),
                    "username": username or "",
                    "full_name": full_name or str(c.id),
                    "phone": phone or "",
                    "access_hash": str(access_hash) if access_hash is not None else "",
                })
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to scrape contacts: {str(e)}")
        finally:
            await _safe_disconnect(client)

    return {"members": members, "count": len(members)}


@router.get("/api/v1/scrape/messaged-users")
async def scrape_messaged_users(
    account_id: str,
    user_id: str = Depends(current_user_id),
) -> dict[str, Any]:
    """Scrape all one-on-one DM users of a Telegram account."""
    members = []
    seen_ids: set[str] = set()
    lock = _account_session_lock(account_id)
    async with lock:
        _, client = await _open_account_client_for_user(user_id, account_id)
        try:
            async for dialog in client.iter_dialogs():
                entity = dialog.entity
                if not hasattr(entity, "id"):
                    continue
                if getattr(entity, "bot", False) or getattr(entity, "deleted", False):
                    continue
                uid = str(entity.id)
                if uid in seen_ids:
                    continue
                if hasattr(entity, "first_name") or hasattr(entity, "last_name"):
                    seen_ids.add(uid)
                    first = getattr(entity, "first_name", None) or ""
                    last = getattr(entity, "last_name", None) or ""
                    full_name = f"{first} {last}".strip()
                    username = getattr(entity, "username", None)
                    phone = getattr(entity, "phone", None)
                    access_hash = getattr(entity, "access_hash", None)
                    members.append({
                        "user_id": uid,
                        "username": username or "",
                        "full_name": full_name or uid,
                        "phone": phone or "",
                        "access_hash": str(access_hash) if access_hash is not None else "",
                    })
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to scrape messaged users: {str(e)}")
        finally:
            await _safe_disconnect(client)

    return {"members": members, "count": len(members)}


@router.get("/api/v1/scrape-group/members/excel")
async def scrape_group_members_excel(
    account_id: str,
    group_id: str,
    user_id: str = Depends(current_user_id),
) -> Response:
    """Scrape members from a group and return as Excel file."""
    members = []
    lock = _account_session_lock(account_id)
    async with lock:
        _, client = await _open_account_client_for_user(user_id, account_id)
        try:
            try:
                entity = await _resolve_entity(client, group_id)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Could not resolve group: {str(e)}")

            try:
                async for participant in client.iter_participants(entity):
                    if getattr(participant, "bot", False) or getattr(participant, "deleted", False):
                        continue
                    username = getattr(participant, "username", None)
                    first_name = getattr(participant, "first_name", None) or ""
                    last_name = getattr(participant, "last_name", None) or ""
                    full_name = f"{first_name} {last_name}".strip()
                    phone = getattr(participant, "phone", None)

                    members.append({
                        "user_id": str(participant.id),
                        "username": username or "",
                        "full_name": full_name or str(participant.id),
                        "phone": phone or "",
                    })
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Failed to scrape members: {str(e)}")

            wb = Workbook()
            ws = wb.active
            ws.title = "Group Members"

            headers = ["User ID", "Username", "Full Name", "Phone"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            for row, member in enumerate(members, 2):
                ws.cell(row=row, column=1, value=member["user_id"])
                ws.cell(row=row, column=2, value=member["username"])
                ws.cell(row=row, column=3, value=member["full_name"])
                ws.cell(row=row, column=4, value=member["phone"])

            for col in range(1, 5):
                max_length = 0
                column_letter = get_column_letter(col)
                for row in range(1, len(members) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            group_name = getattr(entity, "title", "group")
            safe_group_name = "".join(c for c in group_name if c.isalnum() or c in " _-").strip()
            filename = f"{safe_group_name}_members.xlsx"

            return Response(
                content=excel_buffer.read(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )
        finally:
            await _safe_disconnect(client)
